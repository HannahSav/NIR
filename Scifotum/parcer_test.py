# Created by Hannah at 16.07.2024 20:13
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from scipy import interpolate
import math
import neurokit2 as nk
import openpyxl

def save_data(filename, data):
    print(data)
def plot_columns_data_one_plot(columns_data, time, data_name):
    '''
    Строит все линии ЭМГ
    :param columns_data: dict{'data_name':'List[float]}
    :param data_name: название графа
    :return: None
    '''
    if columns_data:
        x_values = time

        plt.figure(figsize=(12, 6))

        for column_name in columns_data.keys():
            if columns_data[column_name] is not None:
                plt.plot(x_values, columns_data[column_name], marker='o', markersize=3, linestyle='-',
                         label=column_name, alpha = 0.5)

        plt.title('{}. Data Plot'.format(data_name))
        plt.xlabel('X[s]')
        plt.ylabel('Value')
        plt.legend()
        plt.grid(True)
        plt.show()



'''
Creating start graphics
'''

def create_old_graphics(data_file, data_name, column_names):
    data = pd.read_excel(data_file)
    columns_data = {}
    time = data['X[s]'].tolist()
    for column_name in column_names:
        if column_name != 'X[s]':
            columns_data[column_name] = data[column_name].tolist()#[1700:2800:5]
    plot_columns_data_one_plot(columns_data, time, data_name)
    return columns_data

def manual_data_markup(data, time, time_stamp_1, time_stamp_2, naming, filepath):
    '''
    :param data:
    :param time:
    :param time_stamp_1:
    :param time_stamp_2:
    :param naming:
    :param filepath:
    :return:
    '''
    print( data)
    new_data = {}
    start_i = 0
    while time[start_i] < time_stamp_1:
        start_i +=1
    end_i = start_i
    while time[end_i] < time_stamp_2:
        end_i += 1
    time = time[start_i:end_i]

    for col in data.keys():
        data[col] = data[col][start_i:end_i]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row = 1, column = 1, value = "X[s]")
    for idx, key in enumerate(data.keys(), start=2):
        sheet.cell(row=1, column=idx, value = key)

    for i, t in enumerate(time, start=2):
        sheet.cell(row=i, column=1, value=t)
        for j, key in enumerate(data.keys(), start=2):
            sheet.cell(row=i, column=j, value=data[key][i-2])
    workbook.save(filepath)







'''
After Maxim smalltalk
'''
def take_data(data_file, data_name, column_names, men, r, gesture):
    '''
    :param men:
    :param r:
    :param gesture:
    :param data_file: str путь до файла
    :param data_name: str название графика
    :param column_names: List[str] названия колонок
    :return: List[float] Данные после чтения
    '''
    def move_to_isoline(data_column):
        '''
        Функция для сдвига к изолинии
        :param data_column: List[float] изначальные данные ЭМГ
        :return: сдвинутые данные ЭМГ
        '''
        mean = np.mean(data_column)
        for i in range(len(data_column)):
            data_column[i] -= mean
        return data_column

    def choose_best_line():
        num = input("Which line was more important (1-8) :").split()[0]
        while num not in ['1', '2', '3', '4', '5', '6', '7', '8']:
            num = input('Try again: ').split()
            print(num)
        return int(num)

    # def creating_emg_amplitude_abs(data_column):
    #     '''
    #     просто модуль
    #
    #     Функция для выделения амплитуды.Пока неправильная
    #     :param data_column: List[float] изначальные данные
    #     :return: List[float] данные после выделения амплитуды
    #     '''
    #     new_data_column = []
    #     for i in range(len(data_column) - 1):
    #         # new_data_column.append(abs(data_column[i] - data_column[i + 1]))
    #         new_data_column.append(abs(data_column[i]))
    #     return new_data_column
    def creating_emg_amplitude_pos_neg(data_column):
        '''
        Искать минимум в отрицательном и максимум в положительном

        Функция для выделения амплитуды. Через соседей меньше больше
        :param data_column: List[float] изначальные данные
        :param time: List[float] моменты времени
        :return: List[float] data данные после выделения амплитуды
        :return: List[float] time временные точки, после нарезки

        '''
        new_data_column = []
        if_positive = data_column[0]
        i = 0
        prev_best = 0
        best_i = 0
        best = 0
        while i < len(data_column):
            new_data_column.append(abs(data_column[i]))
            if abs(data_column[i]) >= best:
                best = abs(data_column[i])
                best_i = i
            elif data_column[i] * if_positive < 0:
                if_positive *= -1
                step = (new_data_column[best_i] - new_data_column[prev_best]) / (best_i - prev_best)
                for j in range(prev_best + 1, best_i):
                    new_data_column[j] = new_data_column[j - 1] + step
                prev_best = best_i
                best = 0
            i += 1
        print("like made")
        return new_data_column


    # def creating_emg_amplitude_with_diff(data_column):
    #     '''
    #     Разница между макс минами(как creating_emg_amplitude_pos_neg но с разницей)
    #
    #     Функция для выделения амплитуды. Через соседей меньше больше
    #     :param data_column: List[float] изначальные данные
    #     :return: List[float] данные после выделения амплитуды
    #
    #     '''
    #     new_data_column = []
    #     new_data_column.append(data_column[0])
    #     # for i in range(1, len(data_column) - 1):
    #         # if data[i] > data[]
    #     # return new_data_column


    #take data


    data = pd.read_excel(data_file)

    data_list = {}
    for column_name in column_names:
        if column_name != 'X[s]':
            data_list[column_name] = data[column_name].tolist()
    time = data['X[s]'].tolist()
    plot_columns_data_one_plot(data_list, time, "TRY_TO_CUT_MOVE.\n" + data_name)


    # СРЕЗАЕМ нули в начале и в конце записи (не изолиния, а именно нули!!!!)
    def find_zeros(in_data, zeros = 0, key ='Avanti sensor 3 - palmaris longus: EMG 3'):
        '''
        :param key:
        :param in_data: словарь массивов эмг
        :param zeros: 0 для срезания краев записи/5% от наибольшего для срезания состояния покоя
        :return: начало и конец нужного среза
        '''
        start_nozeros = 0
        finish_nozeros = -1
        while abs(in_data[key][start_nozeros]) <= zeros:
            start_nozeros+=1
        while abs(in_data[key][finish_nozeros]) <= zeros:
            finish_nozeros -=1
        return start_nozeros, finish_nozeros


    START_NO_ZEROS, STOP_NO_ZEROS = find_zeros(data_list)
    #for start data, for prepared data
    columns_data = {}
    columns_data_new = {}

    # painting start (before find isolines)
    time = time[START_NO_ZEROS:STOP_NO_ZEROS]
    for column_name in data_list.keys():
        columns_data[column_name] = data_list[column_name][START_NO_ZEROS:STOP_NO_ZEROS]
    # plot_columns_data_one_plot(columns_data, time, "START DATA.\n" + data_name)

    #after found isolines
    for column_name, data in columns_data.items():
        # if column_name !=  'Avanti sensor 1 - brachioradialis: EMG 1':
        #     columns_data_new[column_name] = np.zeros(len(time))
        # else:
        columns_data_new[column_name] = move_to_isoline(data)

        # columns_data_new[column_name] = interpolate.interp1d(columns_data_new[column_name], time[:1500], kind="linear")
    # print(columns_data_new)
    plot_columns_data_one_plot(columns_data_new, time, "AFTER MOVING ISOLINE.\n" + data_name)

    time_stamp_1 = float(input("Enter first timestamp: "))
    time_stamp_2 = float(input("Enter second timestamp: "))
    naming = input("naming: ")
    manual_data_markup(columns_data_new, time, time_stamp_1, time_stamp_2, naming, filepath = f"data/{naming}_{men}_{gesture}_{r}.xlsx")
    return


    # Наиболее значимый датчик (на что ориенируемся?)
    best_line = choose_best_line()

    # УБРАЛА ПОКА ЧТО ПОЗИТИВ
    # for column_name, data in columns_data_new.items():
    #     # if column_name ==  'Avanti sensor 1 - brachioradialis: EMG 1':
    #     columns_data_new[column_name] = creating_emg_amplitude_pos_neg(data)
    # plot_columns_data_one_plot(columns_data_new, time, "AFTER CREATING EMG AMPLITUDE.\n"+data_file)

    def cut_iso(in_data, besti, time):
        besti_key = column_names[besti]
        higher_point = abs(max(in_data[besti_key], key = abs))
        start_cut, end_cut = find_zeros(in_data, higher_point * 0.1, besti_key)
        data_cut = {}
        for col in in_data.keys():
            data_cut[col] = in_data[col][start_cut: end_cut]
        time = time[start_cut:end_cut]
        return data_cut, time

    only_gesture_data, time = cut_iso(columns_data_new, best_line, time)
    plot_columns_data_one_plot(only_gesture_data, time, "TRY_TO_CUT_MOVE.\n" + data_name)


    save_data(data_file, columns_data_new)


if __name__ == '__main__':
    # столбцы из xslx, которые нам надо
    column_names = ['X[s]', 'Avanti sensor 1 - brachioradialis: EMG 1', 'Avanti sensor 2: EMG 2',
                    'Avanti sensor 3 - palmaris longus: EMG 3', 'Avanti sensor 4: EMG 4', 'Avanti sensor 5: EMG 5',
                    'Avanti sensor 6: EMG 6', 'Avanti sensor 7: EMG 7', 'Avanti sensor 8: EMG 8']
    # Путь до файла
    gesture = 1
    mens_files = {
        1: {
            "dmitri": ["1.17", "2.18", "3.19", "4.20"], # 8 и немного 6
            "kate": ["2.3", "3.4", "4.5", "5.6"],  # 5 emg
            "kirill": ["1.84", "1.88", "2.85", "2.89", "3.86", "3.90", "4.87"], #4,5 emg
            "maxim": ["1.37", "2.38", "3.39", "4.40"], # 3 emg
            "nikita": ["1.2", "2.3", "3.4", "4.5"], # 3 emg
        },
        2: {
            "dmitri": ["1.21", "2.22", "3.23", "4.24"],  #4,7
            "kate": ["1.7", "2.8", "3.9", "4.10"],  # emg 6
            "kirill": ["1.91", "2.92", "3.93", "4.94"], #  emg 4 6
            "maxim": ["5.45", "6.46", "7.47", "8.48"], # emg 7 6
            "nikita": ["1.6", "2.7", "3.8", "4.9"],  # emg 6 5 4 7
        },
        3: {
            "dmitri": ["1.25", "2.26", "3.27", "4.28"], # 4       6  5 вкл всегда
            "kate": ["1.11", "2.12", "3.13", "4.14"],  # emg 6
            "kirill": ["1.95", "2.96", "3.97", "4.98", "5.99"], #  4 6 emg   5 hold
            "maxim": ["1.49", "2.50", "3.51", "4.52"], # emg  6  3(strange)   5 hold
            "nikita": ["1.10", "2.11", "3.12", "4.13"], #  emg 6,5,4      3, 7 hold
        },
    }
    for gesture in range(1, 2):
        for men, rep_arr in mens_files[gesture].items():
            if True:#men == "kate":
                for r in rep_arr:
                    filepath = f"resources/{gesture}/{gesture}_gesture_{men}_Plot_and_Store_Rep_{r}.xlsx"
                    # # Которая раньше чето строила
                    # create_old_graphics(data_file=data_file, data_name="Pron'", columns_names)
                    # главная преобразующая и получающая данные функция
                    if men == 'kate':
                        take_data(data_file=filepath, data_name=f"{gesture}_{men}_{r}", column_names = column_names[:-1], men=men, r=r, gesture=gesture)
                    else:
                        take_data(data_file=filepath, data_name=f"{gesture}_{men}_{r}", column_names = column_names[:], men=men, r=r, gesture=gesture)

